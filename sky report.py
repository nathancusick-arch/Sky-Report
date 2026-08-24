import io
import os
import re
import zipfile
from datetime import date, datetime, time
from pathlib import PurePosixPath
from xml.sax.saxutils import escape

import pandas as pd
import streamlit as st
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Sky Report Generator", layout="wide")

RAW_SHEET = "Raw Data UK to check"
TOTAL_SHEET = "total number - no checks needed"
SUMMARY_SHEET = "Summary"
COMCAST_SHEET = "Comcast"
POSTCODES_SHEET = "Postcodes"
MACROS_SHEET = "MACROS"

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
Q = lambda name: f"{{{NS_MAIN}}}{name}"

QUESTION_COLUMNS = [
    "When you go to the site, please confirm if the pub was:",
    "If the pub was open, but you couldn't enter, please select why:",
    "1. If other, please explain why:",
    "Please select the reason why the site was closed:",
    "2. If other, please explain why:",
    "Please select the reason why you think the site was closed down:",
    "If changed use, please describe what was at the location now:",
    "3. If other, please explain why:",
    "Was the premises name the same as provided on the report?",
    "If no, please detail any visible name:",
    "Please select which sport you were sent to audit?",
    "Please detail what competition you were sent to audit?",
    "Please detail the relevant match you were sent to audit:",
    "How many TV screens were visible in the venue?",
    "Were any of the TVs switched on?",
    "Were any of the TVs showing sports or sport news?",
    "Was the relevant match showing on the TV screens?",
    "Were any of the TVs showing a Sky Sports channel?",
    "Which channel was shown?",
    "If other, please detail the channel seen:",
    "If not fully visible, please detail what you could read from the logo that was visible:",
    "Please detail the channel showing the relevant game if it was not on Sky Sports",
    "Which sport was showing on the most screens?",
    "If other, please detail the sport seen:",
    "Please detail the team names visible:",
    "Was there a watermark visible on the screen?",
    "If yes, please select the watermark visible:",
    "If other, please describe the watermark:",
    "Did you see a VCID code on the screen?",
    "If yes, please detail the code you saw:",
    "How many people were watching the TV screens?",
    "Please detail the channel you saw on the TV screens:",
    "Was music (live or radio) being played in the venue?",
    "Did you see any internal or external sports advertising at the premises?",
    "If yes, please describe the advertising you saw and where you saw it.",
    "Were you able to enter and view each room in the venue?",
    "If no, please explain why not:",
    "Please confirm the postcode of the relevant site, as displayed on Google search:",
    "Please confirm whether you saw the relevant game showing in the venue:",
    "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
]

BASE_REQUIRED = [
    "order_internal_id", "client_name", "internal_id", "site_internal_id",
    "responsibility", "site_name", "site_address_1", "site_address_2",
    "site_address_3", "site_post_code", "submitted_date", "approval_date",
    "approved_by_name", "item_to_order", "end_date", "date_of_visit_local",
    "time_of_visit_local", "primary_result", "secondary_result", "tertiary_result",
    "site_code",
]

FIELD_ALIASES = {
    "site_internal_id": {
        "siteinternalid", "siteid", "servelegalsiteid", "slsiteid", "internalsiteid", "internalid"
    },
    "account_id": {
        "accountid", "accountid18", "salesforceid", "salesforceaccountid",
        "sfaccountid", "accountidentifier"
    },
    "city": {"city", "town", "sitecity", "sitetown", "towncity"},
    "postcode": {"postcode", "postalcode", "sitepostcode", "eircode"},
    "region": {"region", "newregion", "regionnumber", "regionno"},
    "territory": {"territory", "territorynumber", "territoryno"},
    "record_id": {"recordid", "recordidentifier", "skyrecordid"},
    "sky_reference_number": {
        "skyreferencenumber", "skyreference", "skyrefno", "skyrefnumber", "skyref",
        "ospreyid", "ospreyidentifier",
    },
    "premise_id": {"premiseid", "premisesid", "premiseidnumber", "premisesidentifier"},
    "pot": {"pot", "potid", "potnumber", "potno"},
}

LOOKUP_FIELDS = [
    "city", "region", "territory", "record_id",
    "sky_reference_number", "premise_id", "pot",
]
CRITICAL_LOOKUP_FIELDS = ["city", "region", "territory", "pot"]


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).replace("\u00a0", " ").strip()


def normalise_header(value):
    return re.sub(r"[^a-z0-9]", "", clean_text(value).lower())


def normalise_key(value):
    # Sky's source process uses EXACT against Salesforce Account IDs.
    return clean_text(value)


def normalise_postcode_key(value):
    return re.sub(r"\s+", "", clean_text(value).upper())


def useful(value):
    return value is not None and clean_text(value) != ""


def smart_value(value):
    text = clean_text(value)
    if text == "":
        return None
    low = text.lower()
    if low == "true":
        return True
    if low == "false":
        return False
    if re.fullmatch(r"-?\d+", text) and len(text.lstrip("-")) <= 10:
        try:
            return int(text)
        except ValueError:
            pass
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return float(text)
        except ValueError:
            pass
    return text


def parse_date(value, label="date"):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Could not read {label}: {text!r}")
    return parsed.date()


def parse_time(value):
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Could not read visit time: {text!r}")


def excel_serial(value):
    epoch = datetime(1899, 12, 30)
    if isinstance(value, time):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime.combine(value, time())
    if isinstance(value, datetime):
        delta = value - epoch
        return delta.days + delta.seconds / 86400 + delta.microseconds / 86400000000
    return value


def read_csv_bytes(data):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return pd.read_csv(io.BytesIO(data), dtype=str, keep_default_na=False, encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error or ValueError("Could not decode CSV")


def validate_export(df):
    missing = [c for c in BASE_REQUIRED + QUESTION_COLUMNS if c not in df.columns]
    if missing:
        preview = ", ".join(missing[:8])
        suffix = "…" if len(missing) > 8 else ""
        raise ValueError(f"The audit export is missing {len(missing)} required column(s): {preview}{suffix}")
    sky = df[df["client_name"].str.strip().str.casefold() == "sky"].copy()
    if "status" in sky.columns:
        sky = sky[sky["status"].str.strip().str.casefold() == "approved"].copy()
    if sky.empty:
        raise ValueError("No approved Sky audits were found in the export.")
    return sky.reset_index(drop=True)


def header_field(value):
    norm = normalise_header(value)
    for field, aliases in FIELD_ALIASES.items():
        if norm in aliases:
            return field
    return None


def detect_reference_table(rows):
    best = None
    for row_index, row in enumerate(rows[:15]):
        mapped = {}
        for col_index, value in enumerate(row):
            field = header_field(value)
            if field and field not in mapped:
                mapped[field] = col_index
        has_key = "account_id" in mapped or "site_internal_id" in mapped or "postcode" in mapped
        score = len(mapped) + (3 if has_key else 0)
        if has_key and score >= 4 and (best is None or score > best[0]):
            best = (score, row_index, mapped)
    return best


def merge_lookup(target, key, record, keep_blank=False):
    if not key:
        return
    existing = target.setdefault(key, {})
    for field, value in record.items():
        if useful(value):
            existing[field] = smart_value(value)
        elif keep_blank:
            existing[field] = None


def process_reference_rows(rows, source_label, maps, recognised):
    detected = detect_reference_table(rows)
    if not detected:
        return
    _, header_row, columns = detected
    recognised.append((source_label, sorted(columns)))
    for row in rows[header_row + 1:]:
        record = {}
        for field, col_index in columns.items():
            if col_index < len(row):
                record[field] = row[col_index]
        merge_lookup(maps["site"], normalise_key(record.get("site_internal_id")), record)
        # Preserve explicit blanks in the exact Account ID table so the latest
        # account master remains authoritative over historic LIVE values.
        merge_lookup(
            maps["account"], normalise_key(record.get("account_id")), record,
            keep_blank=True,
        )
        merge_lookup(maps["postcode"], normalise_postcode_key(record.get("postcode")), record)


def read_reference_files(files):
    maps = {"site": {}, "account": {}, "postcode": {}}
    recognised = []
    ignored = []
    for uploaded in files:
        data = uploaded.getvalue()
        suffix = os.path.splitext(uploaded.name)[1].lower()
        try:
            if suffix == ".csv":
                frame = read_csv_bytes(data)
                rows = [list(frame.columns)] + frame.astype(object).where(frame.notna(), None).values.tolist()
                before = len(recognised)
                process_reference_rows(rows, uploaded.name, maps, recognised)
                if len(recognised) == before:
                    ignored.append(uploaded.name)
            else:
                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_links=False)
                found = False
                for ws in wb.worksheets:
                    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200000), values_only=True))
                    before = len(recognised)
                    process_reference_rows(rows, f"{uploaded.name} — {ws.title}", maps, recognised)
                    found = found or len(recognised) > before
                wb.close()
                if not found:
                    ignored.append(uploaded.name)
        except Exception as exc:
            ignored.append(f"{uploaded.name} ({exc})")
    return maps, recognised, ignored


def workbook_sheet_paths(xlsx_bytes):
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        workbook_xml = etree.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {}
    for rel in rels_xml:
        rel_targets[rel.get("Id")] = rel.get("Target")
    result = {}
    sheets = workbook_xml.find(Q("sheets"))
    for index, sheet in enumerate(sheets if sheets is not None else []):
        rel_id = sheet.get(f"{{{NS_REL}}}id")
        target = rel_targets.get(rel_id, "")
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = str(PurePosixPath("xl") / target)
        result[sheet.get("name")] = (path, index)
    return result


def verify_template(xlsx_bytes, required_sheets, label):
    try:
        paths = workbook_sheet_paths(xlsx_bytes)
    except Exception as exc:
        raise ValueError(f"{label} is not a readable .xlsx workbook: {exc}") from exc
    missing = [name for name in required_sheets if name not in paths]
    if missing:
        raise ValueError(f"{label} is missing sheet(s): {', '.join(missing)}")
    return paths


def update_record(existing, values):
    for field, value in values.items():
        if useful(value):
            existing[field] = value


def read_live_support(live_bytes):
    wb = load_workbook(io.BytesIO(live_bytes), read_only=True, data_only=True, keep_links=False)
    existing_audits = set()
    history_site = {}
    history_account = {}
    total = wb[TOTAL_SHEET]
    for row in total.iter_rows(min_row=3, min_col=1, max_col=27, values_only=True):
        audit_id = normalise_key(row[2])
        if audit_id:
            existing_audits.add(audit_id)
        values = {
            "site_internal_id": row[3], "city": row[10], "postcode": row[11],
            "region": row[21], "territory": row[22], "account_id": row[23],
            "record_id": row[24], "sky_reference_number": row[25], "premise_id": row[26],
        }
        # The historic Pot sits in column AD and is read separately below only when needed.
        site_key = normalise_key(row[3])
        account_key = normalise_key(row[23])
        if site_key:
            update_record(history_site.setdefault(site_key, {}), values)
        if account_key:
            update_record(history_account.setdefault(account_key, {}), values)
    total_max_row = total.max_row

    # Read Pot from AD without loading the whole workbook in editable mode.
    for row in total.iter_rows(min_row=3, min_col=4, max_col=30, values_only=True):
        site_key = normalise_key(row[0])
        account_key = normalise_key(row[20])
        pot = row[26]
        if useful(pot):
            if site_key:
                history_site.setdefault(site_key, {})["pot"] = smart_value(pot)
            if account_key:
                history_account.setdefault(account_key, {})["pot"] = smart_value(pot)

    comcast = {}
    for row in wb[COMCAST_SHEET].iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        visit_date = parse_date(row[0]) if row[0] is not None else None
        if visit_date:
            comcast[visit_date] = row[1]

    postcode_country = {}
    ireland_postcode_country = {}
    for row in wb[POSTCODES_SHEET].iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
        if useful(row[0]) and useful(row[2]):
            postcode_country[clean_text(row[0]).upper()] = row[2]
        if useful(row[4]) and useful(row[6]):
            ireland_postcode_country[clean_text(row[4]).upper()] = row[6]
    wb.close()
    return {
        "existing_audits": existing_audits,
        "history_site": history_site,
        "history_account": history_account,
        "total_max_row": total_max_row,
        "comcast": comcast,
        "postcode_country": postcode_country,
        "ireland_postcode_country": ireland_postcode_country,
    }


GB_POSTCODE = re.compile(r"^[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}$", re.I)
EIRCODE = re.compile(r"^[AC-FHKNPRTV-Y]\d{2}\s*[0-9AC-FHKNPRTV-Y]{4}$", re.I)


def detect_orders(df):
    scores = {}
    counts = {}
    for order, group in df.groupby("order_internal_id", sort=False):
        score = 0
        for postcode in group["site_post_code"]:
            pc = re.sub(r"\s+", "", clean_text(postcode).upper())
            if pc.startswith("BT"):
                score += 4
            elif EIRCODE.fullmatch(pc):
                score += 3
            elif GB_POSTCODE.fullmatch(pc):
                score -= 1
        scores[order] = score
        counts[order] = len(group)
    orders = list(counts)
    if not orders:
        return None, None
    if len(orders) == 1:
        return (None, orders[0]) if scores[orders[0]] > 0 else (orders[0], None)
    ireland = max(orders, key=lambda o: (scores[o], -counts[o]))
    remaining = [o for o in orders if o != ireland]
    uk = max(remaining, key=lambda o: counts[o])
    if scores[ireland] <= scores[uk]:
        uk = max(orders, key=lambda o: counts[o])
        ireland_candidates = [o for o in orders if o != uk]
        ireland = max(ireland_candidates, key=lambda o: scores[o]) if ireland_candidates else None
    return uk, ireland


def combine_lookup(row, support, references):
    site_key = normalise_key(row.get("site_internal_id"))
    account_key = normalise_key(row.get("site_code"))
    postcode_key = normalise_postcode_key(row.get("site_post_code"))
    result = {}
    for source in (
        support["history_site"].get(site_key, {}),
        support["history_account"].get(account_key, {}),
        references["postcode"].get(postcode_key, {}),
        references["site"].get(site_key, {}),
    ):
        update_record(result, source)
    account_reference = references["account"].get(account_key, {})
    for field, value in account_reference.items():
        if useful(value):
            result[field] = value
        else:
            result[field] = None
    result["account_id"] = smart_value(row.get("site_code"))
    return result


def build_records(df, support, references):
    records = []
    duplicate_count = 0
    missing_rows = []
    for _, row_series in df.iterrows():
        row = row_series.to_dict()
        if normalise_key(row["internal_id"]) in support["existing_audits"]:
            duplicate_count += 1
            continue
        lookup = combine_lookup(row, support, references)
        account_reference_match = bool(
            references["account"].get(normalise_key(row.get("site_code")))
        )
        questions = [smart_value(row.get(question)) for question in QUESTION_COLUMNS]
        record = {
            "order": smart_value(row["order_internal_id"]),
            "client": smart_value(row["client_name"]),
            "audit": smart_value(row["internal_id"]),
            "site": smart_value(row["site_internal_id"]),
            "end_date": parse_date(row["end_date"], "end_date"),
            "responsibility": smart_value(row["responsibility"]),
            "name": smart_value(row["site_name"]),
            "address1": smart_value(row["site_address_1"]),
            "address2": smart_value(row["site_address_2"]),
            "address3": smart_value(row["site_address_3"]),
            "city": lookup.get("city"),
            "postcode": smart_value(row["site_post_code"]),
            "submitted": parse_date(row["submitted_date"], "submitted_date"),
            "approved": parse_date(row["approval_date"], "approval_date"),
            "approver": None,
            "item": smart_value(row["item_to_order"]),
            "visit_date": parse_date(row["date_of_visit_local"], "date_of_visit_local"),
            "visit_time": parse_time(row["time_of_visit_local"]),
            "primary": smart_value(row["primary_result"]),
            "secondary": smart_value(row["secondary_result"]),
            "tertiary": smart_value(row["tertiary_result"]),
            "region": lookup.get("region"),
            "territory": lookup.get("territory"),
            "account_id": lookup.get("account_id"),
            "record_id": None,
            "sky_reference_number": lookup.get("sky_reference_number"),
            "premise_id": lookup.get("premise_id"),
            "pot": lookup.get("pot"),
            "questions": questions,
        }
        missing = [field for field in CRITICAL_LOOKUP_FIELDS if not useful(record.get(field))]
        if not account_reference_match:
            missing.insert(0, "account reference match")
        if missing:
            missing_rows.append({
                "internal_id": record["audit"], "site_internal_id": record["site"],
                "account_id": record["account_id"], "site_name": record["name"],
                "postcode": record["postcode"], "missing_fields": ", ".join(missing),
            })
        records.append(record)
    return records, duplicate_count, pd.DataFrame(missing_rows)


def postcode_prefix(postcode):
    return clean_text(postcode).upper()[:2]


def country_for(record, uk_order, ireland_order, support):
    prefix = postcode_prefix(record["postcode"])
    if record["order"] == ireland_order:
        return support["ireland_postcode_country"].get(prefix, "ROI")
    if record["order"] == uk_order:
        return support["postcode_country"].get(prefix, "")
    return "Northern Ireland" if prefix == "BT" else ""


def proper_case(value):
    text = clean_text(value)
    return text.title() if text else ""


def dash_if_blank(value):
    return value if useful(value) or isinstance(value, bool) else "-"


def summary_values(record, uk_order, ireland_order, support):
    q = record["questions"]
    first_questions = [dash_if_blank(v) for v in q[10:13]]
    remaining_questions = [dash_if_blank(v) for v in q[0:10] + q[13:40]]
    visit_time = record["visit_time"].strftime("%H:%M") if record["visit_time"] else ""
    return first_questions + [
        country_for(record, uk_order, ireland_order, support),
        support["comcast"].get(record["visit_date"], ""),
        record["pot"], record["region"], record["territory"], record["account_id"],
        record["record_id"] or "", record["sky_reference_number"] if useful(record["sky_reference_number"]) else "",
        record["premise_id"] if useful(record["premise_id"]) else "",
        proper_case(record["name"]), record["address1"] or "", record["address2"] or "",
        record["city"] or "", record["postcode"] or "", record["visit_date"], visit_time,
        record["primary"], record["secondary"], record["tertiary"],
    ] + remaining_questions


def result_bucket(value):
    normalised = re.sub(r"[^a-z]", "", clean_text(value).lower())
    if normalised == "positive":
        return "positive"
    if normalised == "negative":
        return "negative"
    if normalised in {"notaccessed", "abort", "aborted"}:
        return "not_accessed"
    return "other"


def summary_stats(records, order):
    relevant = [record for record in records if record["order"] == order] if order else []
    counts = {"positive": 0, "negative": 0, "not_accessed": 0}
    for record in relevant:
        bucket = result_bucket(record["primary"])
        if bucket in counts:
            counts[bucket] += 1
    counts["total"] = sum(counts.values())
    return counts


def report_basename(records):
    dates = sorted({record["visit_date"] for record in records if record["visit_date"]})
    if not dates:
        return "Sky Report"
    parts = [d.strftime("%d.%m") for d in dates]
    return f"Sky ({', '.join(parts)})"


def row_styles_from_xml(xml_bytes, row_number):
    pattern = rb'<row\b[^>]*\br="' + str(row_number).encode() + rb'"[^>]*>(.*?)</row>'
    match = re.search(pattern, xml_bytes, flags=re.DOTALL)
    styles = {}
    if not match:
        return styles
    for cell in re.finditer(rb'<c\b([^>]*)>', match.group(1)):
        attrs = cell.group(1)
        ref_match = re.search(rb'\br="([A-Z]+)\d+"', attrs)
        style_match = re.search(rb'\bs="(\d+)"', attrs)
        if ref_match and style_match:
            styles[ref_match.group(1).decode()] = style_match.group(1).decode()
    return styles


def xml_safe_text(value):
    text = clean_text(value)
    return "".join(ch for ch in text if ch in "\t\n\r" or ord(ch) >= 32)


def number_text(value):
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (date, datetime, time)):
        value = excel_serial(value)
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value)


def make_cell_xml(ref, value=None, style=None, formula=None, cached=None):
    style_attr = f' s="{style}"' if style is not None else ""
    if formula is not None:
        result = cached
        type_attr = ""
        if isinstance(result, str):
            type_attr = ' t="str"'
        elif isinstance(result, bool):
            type_attr = ' t="b"'
        formula_xml = escape(formula)
        value_xml = "" if result is None else escape(number_text(result))
        return f'<c r="{ref}"{style_attr}{type_attr}><f>{formula_xml}</f><v>{value_xml}</v></c>'
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return f'<c r="{ref}"{style_attr} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float, date, datetime, time)):
        return f'<c r="{ref}"{style_attr}><v>{escape(number_text(value))}</v></c>'
    text = escape(xml_safe_text(value))
    preserve = ' xml:space="preserve"' if text != text.strip() else ""
    return f'<c r="{ref}"{style_attr} t="inlineStr"><is><t{preserve}>{text}</t></is></c>'


def make_row_xml(row_number, values, styles, formulas=None, cached=None, max_col=None):
    formulas = formulas or {}
    cached = cached or {}
    cells = []
    limit = max_col or len(values)
    for index in range(1, limit + 1):
        col = get_column_letter(index)
        ref = f"{col}{row_number}"
        if col in formulas:
            cells.append(make_cell_xml(ref, style=styles.get(col), formula=formulas[col], cached=cached.get(col)))
        else:
            value = values[index - 1] if index - 1 < len(values) else None
            cell = make_cell_xml(ref, value=value, style=styles.get(col))
            if cell:
                cells.append(cell)
    return f'<row r="{row_number}" spans="1:{limit}">{"".join(cells)}</row>'


def total_values(record):
    return [
        record["order"], record["client"], record["audit"], record["site"], record["end_date"],
        record["responsibility"], record["name"], record["address1"], record["address2"],
        record["address3"], record["city"], record["postcode"], record["submitted"],
        record["approved"], record["item"], record["visit_date"], record["visit_time"], None,
        record["primary"], record["secondary"], record["tertiary"], record["region"],
        record["territory"], record["account_id"], record["record_id"],
        record["sky_reference_number"], record["premise_id"], "ALL GOOD", None, record["pot"],
    ] + record["questions"] + [None] * 5


def raw_values(record):
    return [
        None, None, record["order"], record["client"], record["audit"], record["site"],
        record["end_date"], record["responsibility"], record["name"], record["address1"],
        record["address2"], record["address3"], record["city"], record["postcode"], record["submitted"],
        record["approved"], record["item"], record["visit_date"], record["visit_time"], None,
        record["primary"], record["secondary"], record["tertiary"], record["region"],
        record["territory"], record["account_id"], record["record_id"],
        record["sky_reference_number"], record["premise_id"], "ALL GOOD", None, record["pot"],
    ] + record["questions"] + [None] * 4 + [postcode_prefix(record["postcode"])] + [None] * 16


def set_or_create_cell(row, col, row_number, value=None, formula=None, cached=None, style=None):
    ref = f"{col}{row_number}"
    cell = next((c for c in row.findall(Q("c")) if c.get("r") == ref), None)
    if cell is None:
        cell = etree.Element(Q("c"), r=ref)
        row.append(cell)
    existing_style = cell.get("s")
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    cell.attrib.pop("s", None)
    if style is not None:
        cell.set("s", str(style))
    elif existing_style is not None:
        cell.set("s", existing_style)
    if formula is not None:
        f = etree.SubElement(cell, Q("f"))
        f.text = formula
        if isinstance(cached, str):
            cell.set("t", "str")
        elif isinstance(cached, bool):
            cell.set("t", "b")
        v = etree.SubElement(cell, Q("v"))
        v.text = "" if cached is None else number_text(cached)
    elif value is None or value == "":
        row.remove(cell)
        sort_row_cells(row)
        return
    elif isinstance(value, bool):
        cell.set("t", "b")
        etree.SubElement(cell, Q("v")).text = "1" if value else "0"
    elif isinstance(value, (int, float, date, datetime, time)):
        etree.SubElement(cell, Q("v")).text = number_text(value)
    else:
        cell.set("t", "inlineStr")
        inline = etree.SubElement(cell, Q("is"))
        text = etree.SubElement(inline, Q("t"))
        safe = xml_safe_text(value)
        if safe != safe.strip():
            text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        text.text = safe
    sort_row_cells(row)


def sort_row_cells(row):
    cells = list(row.findall(Q("c")))
    if len(cells) < 2:
        return
    for cell in cells:
        row.remove(cell)
    cells.sort(
        key=lambda cell: (
            sum(
                (ord(char) - ord("A") + 1) * (26 ** power)
                for power, char in enumerate(
                    reversed(re.match(r"[A-Z]+", cell.get("r", "A")).group(0))
                )
            )
        )
    )
    for cell in cells:
        row.append(cell)


def replace_data_rows(xml_bytes, start_row, row_xml_strings, end_col):
    root = etree.fromstring(xml_bytes)
    sheet_data = root.find(Q("sheetData"))
    if sheet_data is None:
        raise ValueError("Worksheet has no sheetData element")
    for row in list(sheet_data):
        if int(row.get("r", "0")) >= start_row:
            sheet_data.remove(row)
    for row_xml in row_xml_strings:
        sheet_data.append(etree.fromstring(row_xml.encode("utf-8")))
    last_row = start_row + len(row_xml_strings) - 1
    dimension = root.find(Q("dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:{end_col}{max(last_row, start_row - 1)}")
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def update_raw_xml(xml_bytes, records):
    styles = row_styles_from_xml(xml_bytes, 3)
    header_styles = row_styles_from_xml(xml_bytes, 2)
    rows = []
    for offset, record in enumerate(records):
        row_number = 3 + offset
        values = raw_values(record)
        formulas = {"BY": f"LEFT(N{row_number},2)"}
        cached = {"BY": postcode_prefix(record["postcode"])}
        rows.append(make_row_xml(row_number, values, styles, formulas, cached, max_col=93))
    updated = replace_data_rows(xml_bytes, 3, rows, "CO")
    root = etree.fromstring(updated)
    sheet_data = root.find(Q("sheetData"))
    row1 = next((r for r in sheet_data if r.get("r") == "1"), None)
    row2 = next((r for r in sheet_data if r.get("r") == "2"), None)
    if row1 is not None:
        set_or_create_cell(row1, "C", 1, formula=f"COUNTA(C3:C{len(records)+2})", cached=len(records))
    if row2 is not None:
        # Repair three displaced/missing headers in the supplied working template.
        set_or_create_cell(row2, "AP", 2, value=QUESTION_COLUMNS[9], style=header_styles.get("AP") or header_styles.get("AO"))
        set_or_create_cell(row2, "AT", 2, value=QUESTION_COLUMNS[13], style=header_styles.get("AT") or header_styles.get("AS"))
        set_or_create_cell(row2, "BT", 2, value=QUESTION_COLUMNS[39], style=header_styles.get("BT") or header_styles.get("BS"))
        existing_bu = next((c for c in row2.findall(Q("c")) if c.get("r") == "BU2"), None)
        if existing_bu is not None:
            row2.remove(existing_bu)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def update_macros_xml(xml_bytes, record_count):
    root = etree.fromstring(xml_bytes)
    sheet_data = root.find(Q("sheetData"))
    row3 = next((row for row in sheet_data if row.get("r") == "3"), None)
    if row3 is not None:
        set_or_create_cell(row3, "A", 3, formula=f"'{RAW_SHEET}'!C1", cached=record_count)
        set_or_create_cell(row3, "B", 3, formula="SUM(Summary!B8,Summary!D8)", cached=record_count)
        set_or_create_cell(row3, "C", 3, formula=f"COUNTA(Summary!A12:A{record_count + 11})", cached=record_count)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def summary_formulas(raw_row, summary_row, uk_order):
    raw = f"'{RAW_SHEET}'"
    uk_literal = str(uk_order or "").replace('"', '""')
    formulas = {
        "A": f'IF({raw}!AQ{raw_row}="","-",{raw}!AQ{raw_row})',
        "B": f'IF({raw}!AR{raw_row}="","-",{raw}!AR{raw_row})',
        "C": f'IF({raw}!AS{raw_row}="","-",{raw}!AS{raw_row})',
        "D": f'IF({raw}!C{raw_row}="{uk_literal}",IFERROR(INDEX(Postcodes!$C$2:$C$196,MATCH({raw}!BY{raw_row},Postcodes!$A$2:$A$196,0)),""),IF({raw}!BY{raw_row}="BT","Northern Ireland","ROI"))',
        "E": f'IFERROR(INDEX(Comcast!$B$2:$B$4019,MATCH(R{summary_row},Comcast!$A$2:$A$4019,0)),"")',
        "F": f"{raw}!AF{raw_row}", "G": f"{raw}!X{raw_row}", "H": f"{raw}!Y{raw_row}",
        "I": f"{raw}!Z{raw_row}", "J": f'IF({raw}!AA{raw_row}<>"",{raw}!AA{raw_row},"")',
        "K": f'IF({raw}!AB{raw_row}<>"",{raw}!AB{raw_row},"")',
        "L": f'IF({raw}!AC{raw_row}<>"",{raw}!AC{raw_row},"")',
        "M": f"PROPER({raw}!I{raw_row})", "N": f'IF({raw}!I{raw_row}="","",{raw}!J{raw_row})',
        "O": f"T({raw}!K{raw_row})", "P": f"{raw}!M{raw_row}", "Q": f"{raw}!N{raw_row}",
        "R": f"{raw}!R{raw_row}", "S": f'TEXT({raw}!S{raw_row},"HH:MM")',
        "T": f"{raw}!U{raw_row}", "U": f"{raw}!V{raw_row}", "V": f"{raw}!W{raw_row}",
    }
    raw_question_cols = [get_column_letter(i) for i in range(33, 73)]
    summary_question_cols = [get_column_letter(i) for i in list(range(23, 33)) + list(range(33, 60))]
    selected_raw_cols = raw_question_cols[0:10] + raw_question_cols[13:40]
    for summary_col, raw_col in zip(summary_question_cols, selected_raw_cols):
        formulas[summary_col] = f'IF({raw}!{raw_col}{raw_row}="","-",{raw}!{raw_col}{raw_row})'
    return formulas


def update_summary_xml(xml_bytes, records, uk_order, ireland_order, support, formulas_live):
    styles = row_styles_from_xml(xml_bytes, 12)
    row_xmls = []
    for offset, record in enumerate(records):
        summary_row = 12 + offset
        raw_row = 3 + offset
        values = summary_values(record, uk_order, ireland_order, support)
        formulas = summary_formulas(raw_row, summary_row, uk_order) if formulas_live else {}
        cached = {get_column_letter(i + 1): value for i, value in enumerate(values)}
        row_xmls.append(make_row_xml(summary_row, values, styles, formulas, cached, max_col=59))
    updated = replace_data_rows(xml_bytes, 12, row_xmls, "BK")
    root = etree.fromstring(updated)
    sheet_data = root.find(Q("sheetData"))
    uk_stats = summary_stats(records, uk_order)
    ie_stats = summary_stats(records, ireland_order)
    stats_by_col = {"B": uk_stats, "D": ie_stats}
    for col, stats in stats_by_col.items():
        for row_number, key in ((5, "positive"), (6, "negative"), (7, "not_accessed"), (8, "total")):
            row = next(r for r in sheet_data if r.get("r") == str(row_number))
            value = stats[key]
            if formulas_live:
                if row_number == 8:
                    formula = f"SUM({col}5:{col}7)"
                else:
                    result_text = {5: "Positive", 6: "Negative", 7: "Not Accessed"}[row_number]
                    order = uk_order if col == "B" else ireland_order
                    order_literal = str(order or "").replace('"', '""')
                    formula = f'COUNTIFS(\'{RAW_SHEET}\'!$U$3:$U${len(records)+2},"{result_text}",\'{RAW_SHEET}\'!$C$3:$C${len(records)+2},"{order_literal}")'
                set_or_create_cell(row, col, row_number, formula=formula, cached=value)
            else:
                set_or_create_cell(row, col, row_number, value=value)
        pct_col = "C" if col == "B" else "E"
        for row_number, key in ((5, "positive"), (6, "negative"), (7, "not_accessed")):
            row = next(r for r in sheet_data if r.get("r") == str(row_number))
            pct = stats[key] / stats["total"] if stats["total"] else ""
            if formulas_live:
                formula = f'IFERROR({col}{row_number}/${col}$8,"")'
                set_or_create_cell(row, pct_col, row_number, formula=formula, cached=pct)
            else:
                set_or_create_cell(row, pct_col, row_number, value=pct)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def find_separator_row_template(xml_bytes):
    candidates = re.finditer(rb'<row\b[^>]*>.*?</row>', xml_bytes, flags=re.DOTALL)
    result = None
    for match in candidates:
        row_xml = match.group(0)
        if (
            b'customFormat="1"' in row_xml
            and b"<v>" not in row_xml
            and b"<is>" not in row_xml
            and b"<f" not in row_xml
        ):
            result = row_xml
    if result is None:
        raise ValueError("Could not locate the yellow separator-row style in the historical data sheet.")
    return result


def renumber_row_xml(row_xml, row_number):
    def replace_ref(match):
        column = match.group(1) or b""
        return b'r="' + column + str(row_number).encode() + b'"'
    return re.sub(rb'r="([A-Z]+)?\d+"', replace_ref, row_xml)


def update_total_xml(xml_bytes, records, first_row):
    if not records:
        return xml_bytes
    styles = row_styles_from_xml(xml_bytes, 3)
    separator_template = find_separator_row_template(xml_bytes)
    separator_rows = [
        renumber_row_xml(separator_template, first_row + offset)
        for offset in range(3)
    ]
    data_first_row = first_row + 3
    appended = []
    for offset, record in enumerate(records):
        appended.append(make_row_xml(data_first_row + offset, total_values(record), styles, max_col=75))
    marker = b"</sheetData>"
    position = xml_bytes.find(marker)
    if position < 0:
        raise ValueError("Could not locate sheetData in the historical data sheet.")
    insertion = b"".join(separator_rows) + "".join(appended).encode("utf-8")
    updated = xml_bytes[:position] + insertion + xml_bytes[position:]
    new_last = data_first_row + len(records) - 1
    updated = re.sub(
        rb'(<dimension\b[^>]*\bref=")[^"]+("[^>]*/?>)',
        lambda m: m.group(1) + f"A1:BW{new_last}".encode() + m.group(2),
        updated,
        count=1,
    )
    return updated


def clean_workbook_xml(xml_bytes, active_tab, record_count):
    root = etree.fromstring(xml_bytes)
    external_refs = root.find(Q("externalReferences"))
    if external_refs is not None:
        root.remove(external_refs)
    calc_pr = root.find(Q("calcPr"))
    if calc_pr is None:
        calc_pr = etree.SubElement(root, Q("calcPr"))
    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    book_views = root.find(Q("bookViews"))
    if book_views is not None and len(book_views):
        book_views[0].set("activeTab", str(active_tab))
    defined_names = root.find(Q("definedNames"))
    if defined_names is not None:
        for defined_name in list(defined_names):
            formula = defined_name.text or ""
            if "[" in formula or "#REF!" in formula:
                defined_names.remove(defined_name)
                continue
            if defined_name.get("name") == "_xlnm._FilterDatabase":
                if formula.startswith(f"'{RAW_SHEET}'!"):
                    defined_name.text = f"'{RAW_SHEET}'!$A$2:$CO${record_count + 2}"
                elif formula.startswith(f"{SUMMARY_SHEET}!"):
                    defined_name.text = f"{SUMMARY_SHEET}!$A$11:$AY${record_count + 11}"
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def clean_workbook_rels(xml_bytes):
    root = etree.fromstring(xml_bytes)
    for rel in list(root):
        rel_type = rel.get("Type", "")
        if rel_type.endswith("/externalLink") or rel_type.endswith("/calcChain"):
            root.remove(rel)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def clean_content_types(xml_bytes):
    root = etree.fromstring(xml_bytes)
    for child in list(root):
        part = child.get("PartName", "")
        if part == "/xl/calcChain.xml" or part.startswith("/xl/externalLinks/"):
            root.remove(child)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def remove_external_formulas(xml_bytes):
    return re.sub(rb'<f(?:\s[^>]*)?>[^<]*\[[^<]*</f>', b"", xml_bytes)


def rebuild_package(template_bytes, replacements, active_tab, record_count):
    source = io.BytesIO(template_bytes)
    output = io.BytesIO()
    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zout:
        for info in zin.infolist():
            name = info.filename
            if name == "xl/calcChain.xml" or name.startswith("xl/externalLinks/"):
                continue
            data = zin.read(name)
            if name in replacements:
                data = replacements[name]
            elif name == "xl/workbook.xml":
                data = clean_workbook_xml(data, active_tab, record_count)
            elif name == "xl/_rels/workbook.xml.rels":
                data = clean_workbook_rels(data)
            elif name == "[Content_Types].xml":
                data = clean_content_types(data)
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                data = remove_external_formulas(data)
            zout.writestr(info, data)
    return output.getvalue()


def generate_report(export_df, live_template, reference_files, uk_order, ireland_order):
    live_paths = verify_template(
        live_template,
        [MACROS_SHEET, RAW_SHEET, TOTAL_SHEET, SUMMARY_SHEET, COMCAST_SHEET, POSTCODES_SHEET],
        "Previous LIVE report",
    )
    support = read_live_support(live_template)
    references, recognised, ignored = read_reference_files(reference_files)
    if not references["account"]:
        raise ValueError(
            "The combined Sky account reference file was not recognised. Upload the latest "
            "workbook containing Account ID, NewRegion, Territory and Pot ID columns."
        )
    records, duplicate_count, missing = build_records(export_df, support, references)
    if not records:
        raise ValueError("Every audit in the export is already present in the previous LIVE report.")
    record_orders = {record["order"] for record in records}
    assigned = {order for order in (uk_order, ireland_order) if order}
    unassigned = sorted(record_orders - assigned)
    if unassigned:
        raise ValueError(f"The following order ID(s) have not been assigned to a market: {', '.join(map(str, unassigned))}")
    if not missing.empty:
        return None, records, duplicate_count, missing, recognised, ignored

    with zipfile.ZipFile(io.BytesIO(live_template), "r") as zf:
        total_path, _ = live_paths[TOTAL_SHEET]
        raw_path, raw_index = live_paths[RAW_SHEET]
        summary_path, _ = live_paths[SUMMARY_SHEET]
        macros_path, _ = live_paths[MACROS_SHEET]
        total_xml = update_total_xml(zf.read(total_path), records, support["total_max_row"] + 1)
        raw_xml = update_raw_xml(zf.read(raw_path), records)
        summary_xml = update_summary_xml(zf.read(summary_path), records, uk_order, ireland_order, support, True)
        macros_xml = update_macros_xml(zf.read(macros_path), len(records))
    live_output = rebuild_package(
        live_template,
        {total_path: total_xml, raw_path: raw_xml, summary_path: summary_xml, macros_path: macros_xml},
        raw_index,
        len(records),
    )

    return live_output, records, duplicate_count, missing, recognised, ignored


st.title("Sky Report Generator")
st.write(
    "Generate one combined Sky LIVE report from an approved audit export. "
    "The report can contain England, Scotland, Wales, Northern Ireland and Republic of Ireland visits together."
)

left, right = st.columns(2)
with left:
    export_file = st.file_uploader("1. Approved Sky audit export", type=["csv"])
    live_file = st.file_uploader("2. Most recent Sky LIVE report", type=["xlsx"])
with right:
    sites_file = st.file_uploader(
        "3. Sky sites export",
        type=["csv"],
        help="The sites export must contain internal_id and city columns.",
    )
    reference_file = st.file_uploader(
        "4. Latest Serve Legal UK/ROI combined account reference",
        type=["xlsx", "xlsm"],
        help=(
            "Upload the latest combined workbook with Account ID, OspreyId, PremiseId, "
            "NewRegion, Territory and Pot ID columns."
        ),
    )

export_df = None
uk_order = ireland_order = None
if export_file is not None:
    try:
        export_df = validate_export(read_csv_bytes(export_file.getvalue()))
        uk_order, ireland_order = detect_orders(export_df)
        order_count = export_df["order_internal_id"].nunique()
        st.success(
            f"Found {len(export_df):,} approved Sky audits across "
            f"{export_df['date_of_visit_local'].nunique()} visit date(s) and {order_count} order ID(s). "
            "All rows will be included in one LIVE report."
        )
    except Exception as exc:
        st.error(str(exc))
        export_df = None

ready = all([export_file, live_file, sites_file, reference_file])
if st.button("Generate Sky LIVE report", type="primary", disabled=not ready):
    if export_df is None:
        st.error("Please correct the audit export first.")
    else:
        try:
            lookup_files = [reference_file, sites_file]
            with st.spinner("Building and validating the combined Sky LIVE report…"):
                result = generate_report(
                    export_df,
                    live_file.getvalue(),
                    lookup_files,
                    uk_order,
                    ireland_order,
                )
            live_output, records, duplicate_count, missing, recognised, ignored = result
            if live_output is None:
                st.error(
                    f"Generation stopped because {len(missing):,} audit(s) could not be matched to all required lookup data. "
                    "Check the combined account reference and Sky sites export, or review the diagnostic CSV below."
                )
                st.dataframe(missing.head(100), use_container_width=True, hide_index=True)
                st.download_button(
                    "Download missing lookup diagnostic",
                    missing.to_csv(index=False).encode("utf-8-sig"),
                    "Sky Missing Lookup Values.csv",
                    "text/csv",
                )
            else:
                live_name = f"{report_basename(records)} LIVE.xlsx"
                st.success(
                    f"Generated one combined LIVE report containing {len(records):,} row(s)."
                    + (f" Excluded {duplicate_count:,} audit(s) already in the LIVE history." if duplicate_count else "")
                )
                st.download_button(
                    "Download Sky LIVE report",
                    live_output,
                    live_name,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                with st.expander("Reference-file diagnostics"):
                    if recognised:
                        for label, fields in recognised:
                            st.write(f"✓ {label}: {', '.join(fields)}")
                    else:
                        st.write("No uploaded reference table was recognised.")
                    for item in ignored:
                        st.write(f"Not recognised: {item}")
        except Exception as exc:
            st.exception(exc)
